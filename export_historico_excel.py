#!/usr/bin/env python3
"""
Exporta lotes a Excel usando un archivo de IDs como fuente de verdad.
Por cada ID toma el mejor registro de la BD (mayor precio_total).
Uso: python export_historico_excel.py ids_2026-06-17.txt
     python export_historico_excel.py ids_2026-06-17.txt subasta_junio.xlsx
"""
import sys, io, sqlite3, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import DB_PATH


def exportar(ids_file, output_file=None):
    if not os.path.exists(ids_file):
        print(f"ERROR: No se encontro el archivo {ids_file}")
        sys.exit(1)

    with open(ids_file, encoding='utf-8') as f:
        target_ids = [x.strip() for x in f if x.strip()]

    print(f"IDs a exportar: {len(target_ids)}")

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    placeholders = ','.join('?' * len(target_ids))
    todos = conn.execute(f"""
        SELECT listing_id, titulo, fabricante, modelo, capacidad, grado,
               carrier_lock, cantidad_total, precio_total, precio_unitario_promedio,
               fecha_subasta, dia_semana, hora_cierre, numero_pujas,
               precio_inicio, estado, url
        FROM subastas
        WHERE listing_id IN ({placeholders})
        ORDER BY precio_total DESC NULLS LAST
    """, target_ids).fetchall()

    # Cargar desglose de capacidades desde lote_items para lotes MIXTO
    items_raw = conn.execute(f"""
        SELECT listing_id, capacidad, SUM(cantidad) as cantidad
        FROM lote_items
        WHERE listing_id IN ({placeholders})
        GROUP BY listing_id, capacidad
        ORDER BY listing_id, cantidad DESC
    """, target_ids).fetchall()

    # Diccionario: listing_id -> "128GB(11) / 256GB(11) / 512GB(4)"
    items_por_lote = {}
    for row in items_raw:
        lid = row["listing_id"]
        if lid not in items_por_lote:
            items_por_lote[lid] = []
        if row["capacidad"] and row["capacidad"] not in ("N/A", ""):
            items_por_lote[lid].append(f"{row['capacidad']}({row['cantidad']})")

    conn.close()

    # Un registro por listing_id: prioridad CERRADO > ACTIVO, luego mayor precio
    seen = {}
    for r in todos:
        lid = r["listing_id"]
        if lid not in seen:
            seen[lid] = dict(r)
        else:
            prev = seen[lid]
            prev_cerrado = prev["estado"] == "CERRADO"
            new_cerrado  = r["estado"] == "CERRADO"
            prev_precio  = prev["precio_total"] or 0
            new_precio   = r["precio_total"] or 0
            # Preferir CERRADO; si ambos igual estado, preferir mayor precio
            if (not prev_cerrado and new_cerrado) or \
               (prev_cerrado == new_cerrado and new_precio > prev_precio):
                seen[lid] = dict(r)

    # IDs que no están en la BD (sin captura)
    sin_bd = [lid for lid in target_ids if lid not in seen]
    if sin_bd:
        print(f"  IDs sin datos en BD: {len(sin_bd)}")
        for lid in sin_bd[:5]:
            print(f"    {lid}")

    # Ordenar: modelo > capacidad > grado
    rows = sorted(seen.values(),
                  key=lambda r: (r["modelo"] or "", r["capacidad"] or "", r["grado"] or ""))

    print(f"Lotes encontrados en BD: {len(rows)}")

    # ---- Excel ----
    wb = Workbook()
    ws = wb.active
    ws.title = "Subasta"

    hdr_fill  = PatternFill("solid", start_color="1F4E79")
    hdr_font  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin      = Side(style="thin", color="CCCCCC")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "Listing ID", "LOT # / Titulo", "Fabricante", "Modelo", "Capacidad",
        "Grado", "Carrier", "Unidades", "Precio Total ($)", "Precio/Unidad ($)",
        "Fecha Cierre", "Dia", "Hora Cierre (ET)", "Pujas", "Precio Inicio ($)",
        "Estado", "URL"
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font  = hdr_font
        cell.fill  = hdr_fill
        cell.alignment = hdr_align
        cell.border = border

    ws.row_dimensions[1].height = 30

    alt_fill    = PatternFill("solid", start_color="EBF3FB")
    normal_fill = PatternFill("solid", start_color="FFFFFF")
    data_font   = Font(name="Arial", size=9)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    estado_colores = {
        "CERRADO":  "1D6A1D",
        "ACTIVO":   "2E75B6",
        "RETIRADO": "C00000",
    }

    for row_idx, r in enumerate(rows, 2):
        row_fill = alt_fill if row_idx % 2 == 0 else normal_fill

        precio = r["precio_total"] or 0
        from datetime import date, datetime, timedelta
        hoy = date.today().isoformat()
        fecha_lot = r["fecha_subasta"] or ""
        if r["estado"] == "RETIRADO":
            estado_display = "RETIRADO"
        elif precio > 0 and fecha_lot < hoy:
            estado_display = "CERRADO"
        else:
            estado_display = "ACTIVO"

        # Convertir hora_cierre de UTC a ET (UTC-4) en formato 12h
        hora_et = ""
        if r["hora_cierre"]:
            try:
                dt_utc = datetime.strptime(r["hora_cierre"], "%H:%M:%S")
                dt_et  = dt_utc - timedelta(hours=4)
                hora_et = dt_et.strftime("%I:%M %p").lstrip("0")
            except Exception:
                hora_et = r["hora_cierre"]

        # Capacidad: si es MIXTO y tenemos desglose real, mostrarlo
        cap_raw = r["capacidad"] or "N/A"
        if cap_raw.startswith("MIXTO") and r["listing_id"] in items_por_lote:
            desglose = items_por_lote[r["listing_id"]]
            cap_display = " / ".join(desglose) if desglose else cap_raw
        else:
            cap_display = cap_raw

        values = [
            r["listing_id"],
            r["titulo"],
            r["fabricante"],
            r["modelo"],
            cap_display,
            r["grado"],
            r["carrier_lock"],
            r["cantidad_total"],
            precio,
            r["precio_unitario_promedio"] or 0,
            r["fecha_subasta"],
            r["dia_semana"],
            hora_et,
            r["numero_pujas"],
            r["precio_inicio"] or 0,
            estado_display,
            r["url"],
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font   = data_font
            cell.fill   = row_fill
            cell.border = border
            cell.alignment = left if col in [1, 2, 17] else center

        color = estado_colores.get(estado_display, "555555")
        ws.cell(row=row_idx, column=16).font = Font(name="Arial", size=9, bold=True, color=color)

    widths = [28, 38, 12, 22, 16, 10, 10, 9, 16, 16, 13, 12, 16, 7, 16, 10, 55]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Hoja resumen
    ws2 = wb.create_sheet("Resumen")
    cerrados  = sum(1 for r in rows if (r["precio_total"] or 0) > 0)
    sin_precio = len(rows) - cerrados
    retirados = sum(1 for r in rows if r["estado"] == "RETIRADO")
    total_val = sum(r["precio_total"] or 0 for r in rows)

    resumen = [
        ("Total IDs en archivo",      len(target_ids)),
        ("Encontrados en BD",          len(rows)),
        ("Sin datos en BD",            len(sin_bd)),
        ("Con precio de cierre",       cerrados),
        ("Sin precio (aun abiertos)",  sin_precio - retirados),
        ("Retirados",                  retirados),
        ("Valor total subastado ($)",  round(total_val, 2)),
        ("Archivo IDs origen",         ids_file),
    ]
    for i, (lbl, val) in enumerate(resumen, 1):
        ws2.cell(row=i, column=1, value=lbl).font = Font(bold=True, name="Arial")
        ws2.cell(row=i, column=2, value=val)
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 40

    if not output_file:
        base = os.path.splitext(os.path.basename(ids_file))[0].replace("ids_", "")
        output_file = f"historico_{base}.xlsx"

    wb.save(output_file)
    print(f"Excel guardado: {output_file}")
    print(f"  Con precio : {cerrados}")
    print(f"  Sin precio : {sin_precio - retirados}")
    print(f"  Retirados  : {retirados}")
    print(f"  Valor total: ${total_val:,.2f}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python export_historico_excel.py <ids_file.txt> [output.xlsx]")
        print("  Ejemplo: python export_historico_excel.py ids_2026-06-17.txt")
        sys.exit(1)
    ids_file   = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    exportar(ids_file, output_file)
