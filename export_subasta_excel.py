#!/usr/bin/env python3
"""
Exporta todos los listings de una subasta BStock a Excel.
Lee el token automáticamente de las cookies de Firefox.
Uso: python export_subasta_excel.py
"""
import json
import sys
import os
import shutil
import sqlite3
import requests
from datetime import date, datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

STOREFRONT_ID = "67ec2a5fee190bcb0e7469af"


def leer_cookies_firefox():
    profile_dir = os.path.expandvars(r'%APPDATA%\Mozilla\Firefox\Profiles')
    if not os.path.exists(profile_dir):
        print("❌ No se encontró directorio de perfiles de Firefox")
        return {}
    for p in os.listdir(profile_dir):
        cookies_path = os.path.join(profile_dir, p, 'cookies.sqlite')
        if not os.path.exists(cookies_path):
            continue
        tmp = cookies_path + '.tmp_bstock'
        try:
            shutil.copy2(cookies_path, tmp)
            conn = sqlite3.connect(tmp)
            cur = conn.cursor()
            cur.execute("SELECT name, value FROM moz_cookies WHERE host LIKE '%bstock%'")
            cookies = {name: value for name, value in cur.fetchall()}
            conn.close()
            os.remove(tmp)
            if 'bstock_access_token' in cookies:
                print(f"✅ Token leído de Firefox ({len(cookies['bstock_access_token'])} chars)")
                return cookies
        except Exception as e:
            print(f"⚠️  Error leyendo cookies: {e}")
            if os.path.exists(tmp):
                os.remove(tmp)
    print("❌ No se encontraron cookies de bstock. ¿Estás logueado en Firefox?")
    return {}


def fetch_page(session, token, offset, limit=200):
    payload = {
        "limit": limit,
        "offset": offset,
        "sortBy": "endTime",
        "sortOrder": "asc",
        "storeFrontId": [STOREFRONT_ID]
    }
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
        "Origin": "https://bstock.com",
        "Referer": "https://bstock.com/",
    }
    try:
        r = session.post(
            "https://search.bstock.com/v1/all-listings/listings",
            json=payload, headers=headers, timeout=30
        )
        r.raise_for_status()
        return r.json()
    except Exception as e:
        print(f"❌ Error: {e}")
        return None


def fetch_all(session, token):
    all_listings = []
    offset = 0
    while True:
        print(f"  Descargando lotes {offset}...")
        data = fetch_page(session, token, offset)
        if not data:
            break
        listings = data.get("listings", [])
        total = data.get("total", 0)
        if not listings:
            break
        all_listings.extend(listings)
        offset += len(listings)
        if offset >= total:
            break
    return all_listings


def capacity_str(item):
    caps = item.get("capacity") or item.get("memory") or []
    return ", ".join(caps) if caps else "N/A"


def grade_str(item):
    grades = item.get("sellerGrade") or []
    return ", ".join(grades) if grades else "N/A"


def model_str(item):
    models = item.get("model") or []
    return ", ".join(models) if models else "N/A"


def lot_id_str(item):
    ids = item.get("sellerLotId") or []
    return ", ".join(ids) if ids else ""


def build_excel(listings, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Subasta"

    hdr_fill = PatternFill("solid", start_color="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "LOT #", "Modelo", "Capacidad", "Grado", "Unidades",
        "Precio Total ($)", "Precio/Unidad ($)", "% MSRP",
        "Cierra (UTC)", "Estado", "Deal", "Listing ID", "URL"
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border

    ws.row_dimensions[1].height = 30

    alt_fill = PatternFill("solid", start_color="EBF3FB")
    normal_fill = PatternFill("solid", start_color="FFFFFF")
    data_font = Font(name="Arial", size=9)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    now_utc = datetime.now(timezone.utc)

    for row_idx, item in enumerate(listings, 2):
        winning = item.get("winningBidAmount", 0) or 0
        units = item.get("units", 0) or 0
        pct_msrp = item.get("percentMsrp")
        end_time_raw = item.get("endTime") or ""
        end_time_str = end_time_raw[:16].replace("T", " ")
        deal = item.get("deal") or ""
        url = f"https://bstock.com/buy/listings/details/{item.get('listingId', '')}"

        # Determinar estado real por fecha de cierre UTC
        if end_time_raw:
            try:
                end_dt = datetime.strptime(end_time_raw[:19], '%Y-%m-%dT%H:%M:%S').replace(tzinfo=timezone.utc)
                cerrado = end_dt < now_utc
            except Exception:
                cerrado = winning > 0
        else:
            cerrado = winning > 0

        status = "CERRADO" if cerrado else "ABIERTO"

        # Solo mostrar precio/unidad si realmente cerró
        precio_total = winning if cerrado else 0
        price_per_unit = round(precio_total / units, 2) if units > 0 and precio_total > 0 else 0

        row_fill = alt_fill if row_idx % 2 == 0 else normal_fill

        values = [
            lot_id_str(item), model_str(item), capacity_str(item), grade_str(item),
            units, precio_total, price_per_unit,
            f"{pct_msrp:.1f}%" if pct_msrp else "",
            end_time_str, status, deal, item.get("listingId", ""), url
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = data_font
            cell.fill = row_fill
            cell.border = border
            cell.alignment = center if col not in [1, 2, 13] else left

        status_cell = ws.cell(row=row_idx, column=10)
        if status == "CERRADO":
            status_cell.font = Font(name="Arial", size=9, bold=True, color="1D6A1D")
        else:
            status_cell.font = Font(name="Arial", size=9, bold=True, color="C00000")

        deal_cell = ws.cell(row=row_idx, column=11)
        deal_colors = {"Great Price": "1D6A1D", "Good Price": "2E75B6", "Fair Price": "7F6000"}
        dc = deal_colors.get(deal)
        if dc:
            deal_cell.font = Font(name="Arial", size=9, bold=True, color=dc)

    widths = [22, 22, 18, 10, 9, 16, 16, 9, 16, 9, 12, 28, 50]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    ws2 = wb.create_sheet("Resumen")
    ws2["A1"] = "Total lotes"
    ws2["B1"] = len(listings)
    ws2["A2"] = "Lotes cerrados"
    ws2["B2"] = f"=COUNTIF(Subasta!J2:J{len(listings)+1},\"CERRADO\")"
    ws2["A3"] = "Lotes abiertos"
    ws2["B3"] = f"=COUNTIF(Subasta!J2:J{len(listings)+1},\"ABIERTO\")"
    ws2["A4"] = "Valor total ($)"
    ws2["B4"] = f"=SUM(Subasta!F2:F{len(listings)+1})"
    ws2["A5"] = "Precio/u promedio ($)"
    ws2["B5"] = f"=AVERAGEIF(Subasta!G2:G{len(listings)+1},\">0\")"

    for cell in ["A1", "A2", "A3", "A4", "A5"]:
        ws2[cell].font = Font(bold=True, name="Arial")

    wb.save(filename)
    print(f"✅ Excel guardado: {filename}")


def main():
    cookies = leer_cookies_firefox()
    if not cookies:
        sys.exit(1)

    token = cookies.get('bstock_access_token', '')
    session = requests.Session()
    session.cookies.update(cookies)
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:151.0) Gecko/20100101 Firefox/151.0',
    })

    print("Descargando listings...")
    listings = fetch_all(session, token)
    print(f"Total: {len(listings)} lotes")

    today = date.today().strftime("%Y-%m-%d")
    filename = f"subasta_{today}.xlsx"
    build_excel(listings, filename)
    return filename


if __name__ == "__main__":
    main()
